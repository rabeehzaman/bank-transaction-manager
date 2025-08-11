#!/usr/bin/env python3
import sys
import os
import glob
import re
import csv
import json
import argparse
from typing import Any, List, Tuple, Dict, Optional

try:
    import openpyxl  # type: ignore
except Exception as e:
    print("ERROR: openpyxl is not installed. Run: python3 -m pip install --user openpyxl", file=sys.stderr)
    sys.exit(1)


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\u00a0", " ").replace("\xa0", " ")


def parse_amount(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    s = str(value)
    # Remove SAR symbol \ue900 and other non-numeric characters except . and -
    s = s.replace("\ue900", "").replace(",", "")
    s = re.sub(r"[^0-9.-]", "", s)
    try:
        return float(s)
    except Exception:
        return 0.0


def try_parse_iso(d: Any) -> Optional[str]:
    try:
        # Excel serial
        if isinstance(d, (int, float)):
            from datetime import datetime, timedelta
            # Excel epoch 1900-01-01; adjust for leap bug by -2
            epoch = datetime(1900, 1, 1)
            dt = epoch + timedelta(days=float(d) - 2)
            return dt.date().isoformat()
        # openpyxl already returns datetime/date sometimes
        import datetime as _dt
        if isinstance(d, (_dt.datetime, _dt.date)):
            return (d.date() if isinstance(d, _dt.datetime) else d).isoformat()
        if isinstance(d, str):
            t = d.strip()
            # Prefer DD/MM/YYYY
            m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", t)
            if m:
                day, month, year = map(int, m.groups())
                from datetime import date
                return date(year, month, day).isoformat()
            # Fallback to generic parse
            try:
                from datetime import datetime
                dt = datetime.fromisoformat(t.replace("Z", "+00:00"))
                return dt.date().isoformat()
            except Exception:
                pass
            try:
                import dateutil.parser  # type: ignore
                return dateutil.parser.parse(t).date().isoformat()  # type: ignore
            except Exception:
                # Last resort: simple YYYY-MM-DD pattern
                m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", t)
                if m2:
                    return t
        return None
    except Exception:
        return None


def extract_transactions(raw: List[List[Any]], header_idx: int, cols: Dict[str, Optional[int]]) -> List[Dict[str, Any]]:
    """Extract all transaction-like rows after header using the detected column mapping."""
    def get_num(row: List[Any], idx: Optional[int]) -> float:
        if idx is None or idx >= len(row):
            return 0.0
        return parse_amount(row[idx])

    def get_str(row: List[Any], idx: Optional[int]) -> str:
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    results: List[Dict[str, Any]] = []
    tx_start = header_idx + 1
    for r in raw[tx_start:]:
        if not isinstance(r, list) or len(r) < 2:
            continue
        date_val = r[cols["DATE"]] if cols.get("DATE") is not None and cols["DATE"] < len(r) else None
        date_iso = try_parse_iso(date_val)
        if not date_iso:
            continue
        desc = get_str(r, cols.get("DESCRIPTION"))
        if not desc or try_parse_iso(desc):
            # salvage any non-date text from early columns
            for c in range(min(10, len(r))):
                if c == cols.get("DATE"):
                    continue
                candidate = get_str(r, c)
                if candidate and not try_parse_iso(candidate):
                    desc = candidate
                    break

        ref = get_str(r, cols.get("REFERENCE")) or None

        debit = credit = 0.0
        if cols.get("AMOUNT") is not None and cols.get("DEBIT") is None and cols.get("CREDIT") is None:
            amt = get_num(r, cols["AMOUNT"])  # type: ignore[index]
            credit = amt if amt > 0 else 0.0
            debit = -amt if amt < 0 else 0.0
        else:
            debit = abs(get_num(r, cols.get("DEBIT")))
            credit = get_num(r, cols.get("CREDIT"))

        balance = get_num(r, cols.get("BALANCE"))

        # Skip empties
        if debit == 0 and credit == 0 and not desc:
            continue

        results.append({
            "transaction_date": date_iso,
            "description": desc,
            "reference": ref,
            "debit": debit,
            "credit": credit,
            "balance": balance,
        })

    return results


def detect_header_and_columns(raw: List[List[Any]]) -> Tuple[int, Dict[str, Optional[int]]]:
    best_row = -1
    best_score = -1
    best_cols: Dict[str, Optional[int]] = {}

    for i in range(min(40, len(raw))):
        row = raw[i]
        if not isinstance(row, list):
            continue
        headers = [normalize_header(x) for x in row]

        def find_idx(pred) -> Optional[int]:
            for idx, h in enumerate(headers):
                try:
                    if pred(h):
                        return idx
                except Exception:
                    continue
            return None

        date_idx = find_idx(lambda h: "posting date" in h) or \
                   find_idx(lambda h: "transaction date" in h) or \
                   find_idx(lambda h: h == "date" or h.endswith(" date") or " date" in h)
        desc_idx = find_idx(lambda h: "details" in h or "description" in h or "narration" in h)
        desc_extra_idx = find_idx(lambda h: "description extra" in h or "details extra" in h or "remarks" in h or "note" in h)
        ref_idx = find_idx(lambda h: "reference" in h)
        debit_idx = find_idx(lambda h: "debit" in h or "withdrawal" in h or h == "dr" or h.endswith(" dr"))
        credit_idx = find_idx(lambda h: "credit" in h or "deposit" in h or h == "cr" or h.endswith(" cr"))
        amount_idx = find_idx(lambda h: h == "amount" or h.endswith(" amount") or "amount" in h)
        balance_idx = find_idx(lambda h: "balance" in h)

        present = len([x for x in [date_idx, desc_idx, balance_idx] if x is not None])
        present_money = len([x for x in [debit_idx, credit_idx, amount_idx] if x is not None])
        score = present * 2 + present_money

        if score > best_score:
            best_score = score
            best_row = i
            best_cols = {
                "DATE": date_idx,
                "DESCRIPTION": desc_idx,
                "DESCRIPTION_EXTRA": desc_extra_idx,
                "REFERENCE": ref_idx,
                "DEBIT": debit_idx,
                "CREDIT": credit_idx,
                "AMOUNT": amount_idx,
                "BALANCE": balance_idx,
            }
        if score >= 5:
            break

    if best_row == -1 or not best_cols:
        return 15, {
            "DATE": 0,
            "DESCRIPTION": 1,
            "DESCRIPTION_EXTRA": 2,
            "REFERENCE": 3,
            "DEBIT": 4,
            "CREDIT": 5,
            "AMOUNT": None,
            "BALANCE": 6,
        }
    return best_row, best_cols


def first_non_date_text(row: List[Any], date_col: Optional[int]) -> str:
    for idx, val in enumerate(row[:10]):
        if date_col is not None and idx == date_col:
            continue
        s = str(val or "").strip()
        if not s:
            continue
        if not try_parse_iso(s):
            return s
    return ""


def analyze_sheet(ws) -> Dict[str, Any]:
    # Read a reasonably wide range to capture most structures
    raw: List[List[Any]] = []
    for _r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=500, max_col=100, values_only=True)):
        raw.append(list(row))

    header_idx, cols = detect_header_and_columns(raw)
    # Extract all transactions and sample first 5
    all_tx = extract_transactions(raw, header_idx, cols)
    samples = [
        {
            "date": t["transaction_date"],
            "description": t["description"],
            "reference": t["reference"],
            "debit": t["debit"],
            "credit": t["credit"],
            "balance": t["balance"],
        }
        for t in all_tx[:5]
    ]

    header_values = raw[header_idx] if 0 <= header_idx < len(raw) else []

    # Basic structure metrics
    non_empty_counts_by_col: Dict[int, int] = {}
    non_empty_counts_by_row: Dict[int, int] = {}
    for ri, r in enumerate(raw):
        nnz = 0
        for ci, c in enumerate(r):
            if c not in (None, ""):
                nnz += 1
                non_empty_counts_by_col[ci] = non_empty_counts_by_col.get(ci, 0) + 1
        non_empty_counts_by_row[ri] = nnz

    # Find used range
    used_rows = [ri for ri, cnt in non_empty_counts_by_row.items() if cnt > 0]
    used_cols = [ci for ci, cnt in non_empty_counts_by_col.items() if cnt > 0]
    used_range = {
        "min_row": min(used_rows) + 1 if used_rows else None,
        "max_row": max(used_rows) + 1 if used_rows else None,
        "min_col": min(used_cols) + 1 if used_cols else None,
        "max_col": max(used_cols) + 1 if used_cols else None,
        "non_empty_rows": len(used_rows),
        "non_empty_cols": len(used_cols),
    }

    # Keyword scan (headers/labels anywhere)
    keywords = [
        "date", "transaction date", "posting", "value date", "details", "description",
        "narration", "debit", "credit", "amount", "balance", "reference", "remarks",
        "account", "account number", "opening balance", "closing balance",
    ]
    keyword_hits: Dict[str, List[Tuple[int, int, str]]] = {}
    for ri, r in enumerate(raw):
        for ci, c in enumerate(r):
            if c is None:
                continue
            s = str(c).strip().lower()
            if not s:
                continue
            for kw in keywords:
                if kw in s:
                    keyword_hits.setdefault(kw, []).append((ri + 1, ci + 1, s))

    # Column content typing
    def is_date_like(s: str) -> bool:
        return bool(
            re.match(r"^\d{4}-\d{2}-\d{2}$", s)
            or re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", s)
            or re.match(r"^\d{1,2}-\d{1,2}-\d{4}$", s)
        )

    def is_amount_like(s: str) -> bool:
        s2 = s.replace(",", "")
        return bool(re.match(r"^-?\d+(?:\.\d+)?$", s2))

    column_profiles: Dict[int, Dict[str, Any]] = {}
    for ci in used_cols:
        values = [str(raw[ri][ci]).strip() for ri in used_rows if raw[ri][ci] not in (None, "")]
        if not values:
            continue
        date_cnt = sum(1 for v in values if is_date_like(v) or try_parse_iso(v) is not None)
        amount_cnt = sum(1 for v in values if is_amount_like(v))
        text_cnt = sum(1 for v in values if not (is_date_like(v) or is_amount_like(v)))
        column_profiles[ci + 1] = {
            "non_empty": len(values),
            "date_like": date_cnt,
            "amount_like": amount_cnt,
            "text_like": text_cnt,
            "sample": values[:5],
        }

    # If we couldn't parse any sample rows, prepare a raw preview of the first 30 rows (first 10 columns)
    raw_preview = []
    if len(samples) == 0:
        for idx, r in enumerate(raw[:30], start=1):
            trimmed = []
            for c in r[:10]:
                v = "" if c is None else str(c).strip()
                trimmed.append(v)
            # Skip rows that are entirely empty
            if any(v for v in trimmed):
                raw_preview.append({"row": idx, "cells": trimmed})

    return {
        "header_row": header_idx,
        "columns": cols,
        "header_values": header_values,
        "used_range": used_range,
        "column_profiles": column_profiles,
        "keyword_hits": {k: v[:5] for k, v in keyword_hits.items()},
        "samples": samples,
        "total_transactions": len(all_tx),
        "raw_preview": raw_preview,
        "_all_transactions": all_tx,  # hidden field for export
    }


def analyze_file(path: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        res = analyze_sheet(ws)
        results.append({
            "sheet": sheet_name,
            **res,
            "sample_count": len(res["samples"]),
        })
    return {"file": os.path.basename(path), "sheets": results}


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect and profile GIB Excel statement files")
    parser.add_argument("folder", nargs="?", default=os.path.join(os.getcwd(), "KHALEEJ"), help="Folder containing .xlsx files")
    parser.add_argument("--json", dest="as_json", action="store_true", help="Output JSON summary to stdout")
    parser.add_argument("--export-csv", dest="export_csv", action="store_true", help="Export parsed transactions to CSV per file")
    args = parser.parse_args()

    folder = args.folder
    xlsx_files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
    if not xlsx_files:
        print(f"No .xlsx files found in: {folder}")
        sys.exit(0)

    if not args.as_json:
        print(f"Analyzing {len(xlsx_files)} files in {folder}\n")

    results: List[Dict[str, Any]] = []
    for path in xlsx_files:
        try:
            info = analyze_file(path)
            results.append(info)

            if not args.as_json:
                print(f"File: {info['file']}")
                for sheet in info["sheets"]:
                    print(f"  Sheet: {sheet['sheet']}")
                    print(f"    Header row detected: {sheet['header_row']}")
                    print(f"    Columns: {sheet['columns']}")
                    # Show header values for clarity
                    hv = [str(x) if x is not None else '' for x in sheet.get('header_values', [])]
                    print(f"    Header values: {hv}")
                    print(f"    Used range: {sheet.get('used_range', {})}")
                    print(f"    Column profiles (first 5 samples):")
                    for col_idx, prof in sorted(sheet.get('column_profiles', {}).items()):
                        print(f"      Col {col_idx}: non_empty={prof['non_empty']}, date_like={prof['date_like']}, amount_like={prof['amount_like']}, text_like={prof['text_like']}, sample={prof['sample']}")
                    if sheet.get('keyword_hits'):
                        print("    Keyword hits (up to 5 per keyword):")
                        for kw, hits in sheet['keyword_hits'].items():
                            coords = [f"R{r}C{c}:{txt[:40]}" for (r, c, txt) in hits]
                            print(f"      {kw}: {coords}")
                    print(f"    Sample transactions parsed: {sheet.get('total_transactions', 0)}")
                    for s in sheet["samples"]:
                        print(f"      - {s['date']} | {s['description']} | ref={s['reference']} | dr={s['debit']} | cr={s['credit']} | bal={s['balance']}")
                    if sheet.get('total_transactions', 0) == 0:
                        print("    Raw preview (first 30 rows, 10 cols):")
                        for r in sheet.get('raw_preview', [])[:30]:
                            print(f"      R{r['row']:02d}: {r['cells']}")
                print()
        except Exception as e:
            if not args.as_json:
                print(f"File: {os.path.basename(path)}")
                print(f"  ERROR: {e}")
                print()

    if args.as_json:
        # Remove the heavy _all_transactions when json printing, keep counts
        lightweight = []
        for fobj in results:
            sheets = []
            for s in fobj["sheets"]:
                s_copy = dict(s)
                s_copy.pop("_all_transactions", None)
                sheets.append(s_copy)
            lightweight.append({"file": fobj["file"], "sheets": sheets})
        print(json.dumps(lightweight, ensure_ascii=False, indent=2))

    if args.export_csv:
        out_dir = os.path.join(os.getcwd(), "scripts", "output")
        os.makedirs(out_dir, exist_ok=True)
        for fobj in results:
            file_base = os.path.splitext(fobj["file"])[0]
            for s in fobj["sheets"]:
                tx = s.get("_all_transactions", [])
                if not tx:
                    continue
                out_path = os.path.join(out_dir, f"{file_base} - {s['sheet']}.csv")
                with open(out_path, "w", newline="", encoding="utf-8") as fp:
                    w = csv.DictWriter(fp, fieldnames=["transaction_date", "description", "reference", "debit", "credit", "balance"])
                    w.writeheader()
                    for row in tx:
                        w.writerow(row)
        if not args.as_json:
            print(f"CSV exports written to: {out_dir}")


if __name__ == "__main__":
    main()


